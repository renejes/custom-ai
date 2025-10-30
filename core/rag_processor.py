"""
RAG data processing module for document upload, chunking, and SQLite storage.
Supports PDF, TXT, MD, DOCX, HTML, IPYNB, CSV, XLSX, and JSON files.
"""

import os
import sqlite3
import json
import csv
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from datetime import datetime

# Document loaders
from langchain_text_splitters import RecursiveCharacterTextSplitter

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None

try:
    from docx import Document as DocxDocument
except ImportError:
    DocxDocument = None

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

try:
    import nbformat
except ImportError:
    nbformat = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class DocumentChunk:
    """Represents a chunk of a document."""

    def __init__(self, filename: str, chunk_index: int, text: str, metadata: Dict = None):
        self.filename = filename
        self.chunk_index = chunk_index
        self.text = text
        self.metadata = metadata or {}

    def to_dict(self) -> Dict:
        """Convert to dictionary for database storage."""
        return {
            "filename": self.filename,
            "chunk_index": self.chunk_index,
            "chunk_text": self.text,
            "metadata": json.dumps(self.metadata)
        }


class RAGDatabase:
    """SQLite database for RAG chunks."""

    def __init__(self, db_path: str):
        self.db_path = db_path
        self.conn: Optional[sqlite3.Connection] = None
        self._initialize_db()

    def _initialize_db(self):
        """Create database and table if not exists."""
        self.conn = sqlite3.connect(self.db_path)
        cursor = self.conn.cursor()

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT NOT NULL,
                chunk_index INTEGER NOT NULL,
                chunk_text TEXT NOT NULL,
                metadata TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Create index for faster queries
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_filename
            ON documents(filename)
        """)

        self.conn.commit()

    def insert_chunks(self, chunks: List[DocumentChunk]) -> int:
        """
        Insert multiple chunks into database.

        Args:
            chunks: List of DocumentChunk objects

        Returns:
            Number of chunks inserted
        """
        cursor = self.conn.cursor()

        for chunk in chunks:
            cursor.execute("""
                INSERT INTO documents (filename, chunk_index, chunk_text, metadata)
                VALUES (?, ?, ?, ?)
            """, (
                chunk.filename,
                chunk.chunk_index,
                chunk.text,
                json.dumps(chunk.metadata)
            ))

        self.conn.commit()
        return len(chunks)

    def get_document_count(self) -> int:
        """Get total number of unique documents."""
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(DISTINCT filename) FROM documents")
        return cursor.fetchone()[0]

    def get_chunk_count(self) -> int:
        """Get total number of chunks."""
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM documents")
        return cursor.fetchone()[0]

    def get_processed_files(self) -> List[str]:
        """Get list of already processed filenames."""
        cursor = self.conn.cursor()
        cursor.execute("SELECT DISTINCT filename FROM documents")
        return [row[0] for row in cursor.fetchall()]

    def document_exists(self, filename: str) -> bool:
        """Check if document is already processed."""
        return filename in self.get_processed_files()

    def delete_document(self, filename: str) -> int:
        """
        Delete all chunks of a document.

        Args:
            filename: Name of the document to delete

        Returns:
            Number of chunks deleted
        """
        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM documents WHERE filename = ?", (filename,))
        deleted = cursor.rowcount
        self.conn.commit()
        return deleted

    def close(self):
        """Close database connection."""
        if self.conn:
            self.conn.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


class DocumentLoader:
    """Loads text from various document formats."""

    @staticmethod
    def load_text(file_path: str) -> Optional[str]:
        """
        Load text from a file based on extension.

        Args:
            file_path: Path to the file

        Returns:
            Extracted text or None if failed
        """
        ext = Path(file_path).suffix.lower()

        try:
            if ext in ['.txt', '.md']:
                return DocumentLoader._load_text_file(file_path)
            elif ext == '.pdf':
                return DocumentLoader._load_pdf(file_path)
            elif ext == '.docx':
                return DocumentLoader._load_docx(file_path)
            elif ext == '.html' or ext == '.htm':
                return DocumentLoader._load_html(file_path)
            elif ext == '.ipynb':
                return DocumentLoader._load_notebook(file_path)
            elif ext == '.csv':
                return DocumentLoader._load_csv(file_path)
            elif ext == '.xlsx':
                return DocumentLoader._load_excel(file_path)
            elif ext == '.json':
                return DocumentLoader._load_json(file_path)
            else:
                return None
        except Exception as e:
            print(f"Error loading {file_path}: {e}")
            return None

    @staticmethod
    def _load_text_file(file_path: str) -> str:
        """Load plain text or markdown file."""
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()

    @staticmethod
    def _load_pdf(file_path: str) -> Optional[str]:
        """Load PDF file."""
        if PdfReader is None:
            raise ImportError("pypdf not installed. Run: pip install pypdf")

        reader = PdfReader(file_path)
        text_parts = []

        for page in reader.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)

        return "\n\n".join(text_parts)

    @staticmethod
    def _load_docx(file_path: str) -> Optional[str]:
        """Load DOCX file."""
        if DocxDocument is None:
            raise ImportError("python-docx not installed. Run: pip install python-docx")

        doc = DocxDocument(file_path)
        text_parts = []

        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_parts.append(paragraph.text)

        return "\n\n".join(text_parts)

    @staticmethod
    def _load_html(file_path: str) -> Optional[str]:
        """Load HTML file and extract text."""
        if BeautifulSoup is None:
            raise ImportError("beautifulsoup4 not installed. Run: pip install beautifulsoup4")

        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            html_content = f.read()

        soup = BeautifulSoup(html_content, 'lxml')

        # Remove script and style elements
        for script in soup(["script", "style"]):
            script.decompose()

        # Get text
        text = soup.get_text()

        # Clean up whitespace
        lines = (line.strip() for line in text.splitlines())
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        text = '\n'.join(chunk for chunk in chunks if chunk)

        return text

    @staticmethod
    def _load_notebook(file_path: str) -> Optional[str]:
        """Load Jupyter notebook and extract markdown + outputs."""
        if nbformat is None:
            raise ImportError("nbformat not installed. Run: pip install nbformat")

        with open(file_path, 'r', encoding='utf-8') as f:
            notebook = nbformat.read(f, as_version=4)

        text_parts = []

        for cell in notebook.cells:
            if cell.cell_type == 'markdown':
                # Add markdown content
                text_parts.append(cell.source)
            elif cell.cell_type == 'code':
                # Add code output (text/plain)
                if hasattr(cell, 'outputs'):
                    for output in cell.outputs:
                        if hasattr(output, 'text'):
                            text_parts.append(output.text)
                        elif hasattr(output, 'data') and 'text/plain' in output.data:
                            text_parts.append(output.data['text/plain'])

        return "\n\n".join(text_parts)

    @staticmethod
    def _load_csv(file_path: str) -> str:
        """Load CSV file and convert to text."""
        text_parts = []

        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            for row in reader:
                if row:  # Skip empty rows
                    text_parts.append(" | ".join(row))

        return "\n".join(text_parts)

    @staticmethod
    def _load_excel(file_path: str) -> Optional[str]:
        """Load Excel file and convert to text."""
        if load_workbook is None:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")

        workbook = load_workbook(file_path, data_only=True)
        text_parts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"Sheet: {sheet_name}")

            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    text_parts.append(row_text)

            text_parts.append("")  # Empty line between sheets

        return "\n".join(text_parts)

    @staticmethod
    def _load_json(file_path: str) -> str:
        """Load JSON file and convert to text."""
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Convert JSON to formatted text
        def flatten_json(obj, parent_key='', sep='.'):
            """Flatten nested JSON to readable text."""
            items = []
            if isinstance(obj, dict):
                for k, v in obj.items():
                    new_key = f"{parent_key}{sep}{k}" if parent_key else k
                    if isinstance(v, (dict, list)):
                        items.extend(flatten_json(v, new_key, sep=sep))
                    else:
                        items.append(f"{new_key}: {v}")
            elif isinstance(obj, list):
                for i, item in enumerate(obj):
                    new_key = f"{parent_key}[{i}]"
                    if isinstance(item, (dict, list)):
                        items.extend(flatten_json(item, new_key, sep=sep))
                    else:
                        items.append(f"{new_key}: {item}")
            else:
                items.append(f"{parent_key}: {obj}")
            return items

        lines = flatten_json(data)
        return "\n".join(lines)


class RAGProcessor:
    """Main RAG processing class."""

    def __init__(self, db_path: str, chunk_size: int = 512, chunk_overlap: int = 50):
        """
        Initialize RAG processor.

        Args:
            db_path: Path to SQLite database
            chunk_size: Number of characters per chunk
            chunk_overlap: Number of characters overlap between chunks
        """
        self.db_path = db_path
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap

        # Initialize text splitter
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            length_function=len,
            separators=["\n\n", "\n", ". ", " ", ""]
        )

    def process_file(self, file_path: str, overwrite: bool = False) -> Tuple[bool, str, int]:
        """
        Process a single file: load, chunk, and store in database.

        Args:
            file_path: Path to the file
            overwrite: If True, delete existing chunks before processing

        Returns:
            Tuple of (success: bool, message: str, num_chunks: int)
        """
        filename = os.path.basename(file_path)

        # Check if file exists
        if not os.path.exists(file_path):
            return False, f"File not found: {file_path}", 0

        # Load document
        text = DocumentLoader.load_text(file_path)
        if not text:
            return False, f"Could not extract text from {filename}", 0

        if len(text.strip()) == 0:
            return False, f"File {filename} is empty", 0

        # Check if already processed
        with RAGDatabase(self.db_path) as db:
            if db.document_exists(filename):
                if not overwrite:
                    return False, f"File {filename} already processed (use overwrite=True to reprocess)", 0
                else:
                    db.delete_document(filename)

            # Split into chunks
            chunks = self.text_splitter.split_text(text)

            # Create DocumentChunk objects
            doc_chunks = []
            for idx, chunk_text in enumerate(chunks):
                metadata = {
                    "source": filename,
                    "chunk_index": idx,
                    "total_chunks": len(chunks),
                    "file_type": Path(file_path).suffix,
                    "processed_at": datetime.now().isoformat()
                }

                doc_chunk = DocumentChunk(
                    filename=filename,
                    chunk_index=idx,
                    text=chunk_text,
                    metadata=metadata
                )
                doc_chunks.append(doc_chunk)

            # Insert into database
            num_inserted = db.insert_chunks(doc_chunks)

            return True, f"Successfully processed {filename}: {num_inserted} chunks", num_inserted

    def process_files(self, file_paths: List[str], overwrite: bool = False) -> Dict:
        """
        Process multiple files.

        Args:
            file_paths: List of file paths
            overwrite: If True, delete existing chunks before processing

        Returns:
            Dictionary with processing results
        """
        results = {
            "total_files": len(file_paths),
            "successful": 0,
            "failed": 0,
            "total_chunks": 0,
            "details": []
        }

        for file_path in file_paths:
            success, message, num_chunks = self.process_file(file_path, overwrite)

            if success:
                results["successful"] += 1
                results["total_chunks"] += num_chunks
            else:
                results["failed"] += 1

            results["details"].append({
                "filename": os.path.basename(file_path),
                "success": success,
                "message": message,
                "chunks": num_chunks
            })

        return results

    def get_stats(self) -> Dict:
        """
        Get database statistics.

        Returns:
            Dictionary with stats
        """
        with RAGDatabase(self.db_path) as db:
            return {
                "total_documents": db.get_document_count(),
                "total_chunks": db.get_chunk_count(),
                "processed_files": db.get_processed_files(),
                "database_path": self.db_path,
                "database_size_mb": round(os.path.getsize(self.db_path) / (1024*1024), 2) if os.path.exists(self.db_path) else 0
            }

    def export_database(self) -> str:
        """
        Get database path for export.

        Returns:
            Path to database file
        """
        return self.db_path

    def import_database(self, source_db_path: str, mode: str = "merge") -> Tuple[bool, str, int]:
        """
        Import chunks from another database.

        Args:
            source_db_path: Path to source database
            mode: "merge" (add new chunks) or "replace" (delete existing first)

        Returns:
            Tuple of (success, message, num_chunks_imported)
        """
        if not os.path.exists(source_db_path):
            return False, f"Source database not found: {source_db_path}", 0

        # Connect to source database
        try:
            source_conn = sqlite3.connect(source_db_path)
            source_cursor = source_conn.cursor()

            # Get all chunks from source
            source_cursor.execute("SELECT filename, chunk_index, chunk_text, metadata FROM documents")
            rows = source_cursor.fetchall()
            source_conn.close()

            if not rows:
                return False, "Source database is empty", 0

            # Import into current database
            with RAGDatabase(self.db_path) as db:
                if mode == "replace":
                    # Clear current database
                    db.conn.execute("DELETE FROM documents")
                    db.conn.commit()

                # Track which files already exist
                existing_files = set(db.get_processed_files()) if mode == "merge" else set()

                chunks_imported = 0
                chunks_skipped = 0

                for filename, chunk_index, chunk_text, metadata_str in rows:
                    # Skip if file exists and we're in merge mode
                    if mode == "merge" and filename in existing_files:
                        chunks_skipped += 1
                        continue

                    # Parse metadata
                    try:
                        metadata = json.loads(metadata_str) if metadata_str else {}
                    except:
                        metadata = {}

                    # Create chunk
                    chunk = DocumentChunk(filename, chunk_index, chunk_text, metadata)

                    # Insert
                    db.insert_chunks([chunk])
                    chunks_imported += 1

            message_parts = [f"Successfully imported {chunks_imported} chunks"]
            if chunks_skipped > 0:
                message_parts.append(f"(skipped {chunks_skipped} duplicate chunks)")

            return True, " ".join(message_parts), chunks_imported

        except Exception as e:
            return False, f"Error importing database: {e}", 0

    def export_to_json(self, output_path: str) -> Tuple[bool, str]:
        """
        Export database to JSON format.

        Args:
            output_path: Path for JSON output file

        Returns:
            Tuple of (success, message)
        """
        try:
            with RAGDatabase(self.db_path) as db:
                cursor = db.conn.cursor()
                cursor.execute("SELECT filename, chunk_index, chunk_text, metadata FROM documents ORDER BY filename, chunk_index")
                rows = cursor.fetchall()

                # Group by filename
                documents = {}
                for filename, chunk_index, chunk_text, metadata_str in rows:
                    if filename not in documents:
                        documents[filename] = []

                    try:
                        metadata = json.loads(metadata_str) if metadata_str else {}
                    except:
                        metadata = {}

                    documents[filename].append({
                        "chunk_index": chunk_index,
                        "text": chunk_text,
                        "metadata": metadata
                    })

                # Write to JSON
                output_data = {
                    "export_date": datetime.now().isoformat(),
                    "total_documents": len(documents),
                    "total_chunks": len(rows),
                    "documents": documents
                }

                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(output_data, f, indent=2, ensure_ascii=False)

                return True, f"Exported {len(rows)} chunks to {output_path}"

        except Exception as e:
            return False, f"Error exporting to JSON: {e}"

    def export_to_jsonl(self, output_path: str) -> Tuple[bool, str]:
        """
        Export database to JSONL format (one chunk per line).

        Args:
            output_path: Path for JSONL output file

        Returns:
            Tuple of (success, message)
        """
        try:
            with RAGDatabase(self.db_path) as db:
                cursor = db.conn.cursor()
                cursor.execute("SELECT filename, chunk_index, chunk_text, metadata FROM documents ORDER BY filename, chunk_index")
                rows = cursor.fetchall()

                with open(output_path, 'w', encoding='utf-8') as f:
                    for filename, chunk_index, chunk_text, metadata_str in rows:
                        try:
                            metadata = json.loads(metadata_str) if metadata_str else {}
                        except:
                            metadata = {}

                        chunk_data = {
                            "filename": filename,
                            "chunk_index": chunk_index,
                            "text": chunk_text,
                            "metadata": metadata
                        }

                        f.write(json.dumps(chunk_data, ensure_ascii=False) + '\n')

                return True, f"Exported {len(rows)} chunks to {output_path}"

        except Exception as e:
            return False, f"Error exporting to JSONL: {e}"


if __name__ == "__main__":
    # Test RAG processor
    import tempfile

    # Create test database
    db_path = os.path.join(tempfile.gettempdir(), "test_rag.db")
    processor = RAGProcessor(db_path)

    # Create test file
    test_file = os.path.join(tempfile.gettempdir(), "test.txt")
    with open(test_file, 'w') as f:
        f.write("This is a test document. " * 100)

    # Process file
    success, message, chunks = processor.process_file(test_file)
    print(f"Processing: {message}")

    # Get stats
    stats = processor.get_stats()
    print(f"\nStats: {stats}")

    # Cleanup
    os.remove(test_file)
    os.remove(db_path)
